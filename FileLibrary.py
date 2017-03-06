import openpyxl
from robot.api import logger

def load_file(path):
  file = openpyxl.load_workbook(path)
  sheet = file.active
  return sheet

def get_record_row(hasHeader):
  return 2 if hasHeader else 1

def read_record(sheet, row, keyList, keyDict):
  dictionary = {}
  for key in keyList:
    cell = key + str(row)
    val = sheet[cell].value
    k = keyDict.get(key, None)
    dictionary[k] = str(val)
  return dictionary

def read_row(sheet, row, keyList):
  list = []
  for key in keyList:
    cell = key + str(row)
    val = sheet[cell].value
    list.append(str(val))
  return list

def log(text, debug):
  if debug:
    logger.console(text)

def read_multiple_records(path, keyDict, hasHeader=True, debug=False):
  sheet = load_file(path)
  headerRow = get_record_row(hasHeader)

  keyList = keyDict.keys()
  list = []

  log('', debug)
  for row in range(headerRow, sheet.max_row + 1):
    dictionary = read_record(sheet, row, keyList, keyDict)
    log('[' + str(row) + '] =' + str(dictionary), debug)
    list.append(dictionary)
  log(list, debug)
  return list


def read_single_record(path, keyDict, row='1', debug=False):
  sheet = load_file(path)
  keyList = keyDict.keys()
  list = []
  log('', debug)

  if ( int(row) <= sheet.max_row ):
    dictionary = read_record(sheet, row, keyList, keyDict)
    log('[' + row + '] =' + str(dictionary), debug)
    list.append(dictionary)

  log(list, debug)
  return list


def read_single_row(path, keyDict, row='1', debug=False):
  sheet = load_file(path)
  keyList = keyDict.keys()
  list = []
  log('', debug)

  if ( int(row) <= sheet.max_row ):
    itemList = read_row(sheet, row, keyList)
    log('[' + row + '] =' + str(itemList), debug)
    list = itemList

  log(list, debug)
  return list